# -*- coding: utf-8 -*-
"""lingxing.py — 领星批量导入采购单生成

规则（需求已对齐）:
- 以领星模板为底版, 清掉示例行, 从 R3 起每行一个 SKU
- 标识号(A) = 款×工厂组合顺序 1/2/3...（与订单号序号一致）
- 采购单号(B) = 订单号 PPY+日期+序号
- 供应商(C) = 工厂名 + "工厂" 后缀
- 含税(L)="否", 费用分配方式(M)="不分配", 采购币种(N)="CNY"
- 采购仓库(W) = 配置表「采购仓库」
- SKU(Y) = 主表 SKU编码, 实际采购量(AE) = 采购数
- 其余列一律留空（模板示例留空的字段不填）
"""
import os
from typing import Dict, List, Tuple

import openpyxl

from .parser import InputData
from .splitter import SplitResult

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LX_TEMPLATE = os.path.join(BASE_DIR, "templates", "lingxing_template.xlsx")

# 列索引 (1-based)
COL_ID, COL_ORDER_NO, COL_SUPPLIER, COL_TAX, COL_FEE, COL_CURRENCY, COL_WAREHOUSE, COL_SKU, COL_QTY = \
    1, 2, 3, 12, 13, 14, 23, 25, 31


def build_lingxing(data: InputData, split_result: SplitResult, out_dir: str) -> str:
    wb = openpyxl.load_workbook(LX_TEMPLATE, data_only=False)
    ws = wb["采购单"]

    # 1. 清空数据区（R3 起），保留表头 R2 与格式
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 2. 建立 (款号, 工厂) -> 标识号 映射（与订单号序号一致）
    id_map: Dict[Tuple[str, str], int] = {}
    order_no_map: Dict[Tuple[str, str], str] = {}
    for i, u in enumerate(split_result.units, 1):
        id_map[(u.style, u.factory)] = i
        order_no_map[(u.style, u.factory)] = u.order_no

    # 3. 填数据（每行一个 SKU）
    r = 3
    for sr in data.rows:
        key = (sr.style, sr.factory)
        ws.cell(row=r, column=COL_ID, value=id_map.get(key, 1))
        ws.cell(row=r, column=COL_ORDER_NO, value=order_no_map.get(key, ""))
        ws.cell(row=r, column=COL_SUPPLIER, value=f"{sr.factory}工厂")
        ws.cell(row=r, column=COL_TAX, value="否")
        ws.cell(row=r, column=COL_FEE, value="不分配")
        ws.cell(row=r, column=COL_CURRENCY, value="CNY")
        ws.cell(row=r, column=COL_WAREHOUSE, value=data.config.warehouse)
        ws.cell(row=r, column=COL_SKU, value=sr.sku)
        ws.cell(row=r, column=COL_QTY, value=sr.qty)
        r += 1

    path = os.path.join(out_dir, f"{date_str}_领星批量导入采购单.xlsx"
                         if (date_str := _extract_date(data.input_filename)) else "领星批量导入采购单.xlsx")
    wb.save(path)
    return path


def _extract_date(filename: str) -> str:
    import re
    from datetime import date
    m = re.search(r"(\d{8})", filename)
    return m.group(1) if m else date.today().strftime("%Y%m%d")
