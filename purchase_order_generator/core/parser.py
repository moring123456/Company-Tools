# -*- coding: utf-8 -*-
"""parser.py — 输入解析 + 校验

输入文件: 1 个 .xlsx，3 个工作表
  Sheet1 主表  : 款号|SKU编码|SKU名称|颜色|尺码|采购数|工厂|布行|品名|备注
  Sheet2 面料表: 款号|产品名称|品名|纸样名称|洗水唛成分|每条布出货数(件)|是否需要压缩袋|包装袋规格
  Sheet3 配置表: 品牌 / 账号代码 / 申请人 / 联系地址（A列键 B列值）
"""
from dataclasses import dataclass, field
from typing import Dict, List, Tuple
import re

import openpyxl

from .calc import SIZE_ORDER, is_mixed_color, split_suppliers


# ---------- 数据结构 ----------

@dataclass
class SKURow:
    style: str = ""
    sku: str = ""
    sku_name: str = ""
    color: str = ""
    size: str = ""
    qty: int = 0
    factory: str = ""
    supplier: str = ""       # 布行（混搭色含逗号两个）
    fabric: str = ""         # 品名
    remark: str = ""
    row_no: int = 0          # 输入表行号（报错定位用）


@dataclass
class FabricInfo:
    style: str = ""
    product_name: str = ""
    fabric: str = ""
    pattern: str = ""        # 纸样名称
    wash_label: str = ""     # 洗水唛成分（多行字符串，备用）
    wash_components: List[str] = field(default_factory=list)  # 洗水唛成分列表（每个成分一项，用于公式拼接）
    pcs_per_roll: int = 40   # 每条布出货数
    need_bag: bool = False   # 是否需要压缩袋
    bag_spec: str = ""       # 包装袋规格


@dataclass
class Config:
    brand: str = ""
    account_code: str = ""
    applicant: str = ""
    office: str = ""
    warehouse: str = ""   # 采购仓库（领星批量导入用）


@dataclass
class InputData:
    rows: List[SKURow] = field(default_factory=list)
    fabrics: Dict[str, FabricInfo] = field(default_factory=dict)   # key=款号
    config: Config = field(default_factory=Config)
    images: Dict[str, bytes] = field(default_factory=dict)         # 位置名 → 图片bytes（第4 sheet）
    input_filename: str = ""
    errors: List[Tuple[int, str, str, str]] = field(default_factory=list)  # (行号, 字段, 问题, 建议)


# 图片 sheet 位置约定: 标签行(1-based) → 位置名
IMAGE_LABELS = {
    2: "包装袋图片",
    3: "主唛",
    4: "吊牌",
    5: "温馨提示卡片",
}


# ---------- Sheet 定位 ----------

def _find_sheet(wb, keywords: List[str], fallback_idx: int):
    """按 sheet 名关键词匹配；失败则按位置回退"""
    for ws in wb.worksheets:
        for kw in keywords:
            if kw in ws.title:
                return ws
    if len(wb.worksheets) > fallback_idx:
        return wb.worksheets[fallback_idx]
    return None


# ---------- 主表解析 ----------

MAIN_HEADER_MAP = {
    "款号": "style", "sku编码": "sku", "sku名称": "sku_name", "颜色": "color",
    "尺码": "size", "采购数": "qty", "采购数量": "qty", "采购": "qty",
    "工厂": "factory", "加工厂": "factory", "布行": "supplier",
    "品名": "fabric", "面料名": "fabric", "备注": "remark",
}


def _parse_main(ws, errors: List[Tuple[int, str, str, str]]) -> List[SKURow]:
    rows: List[SKURow] = []
    # 定位表头行（前 10 行内找含"款号"和"SKU"的行）
    header_row, col_map = None, {}
    for r in range(1, min(ws.max_row, 10) + 1):
        row_vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        if any("款号" in v for v in row_vals) and any("SKU" in v.upper() for v in row_vals):
            header_row = r
            col_map = {}
            for idx, v in enumerate(row_vals):
                v_low = v.lower().replace(" ", "")
                for key, field_name in MAIN_HEADER_MAP.items():
                    if key in v_low:
                        col_map[field_name] = idx
                        break
            break
    if header_row is None:
        errors.append((0, "主表", "未找到表头行", "主表需包含 款号/SKU编码/... 等列"))
        return rows
    required = ["style", "sku", "sku_name", "color", "size", "qty", "factory", "supplier", "fabric", "remark"]
    missing = [f for f in required if f not in col_map]
    if missing:
        errors.append((header_row, "主表", f"缺少列: {','.join(missing)}", "表头需 10 列齐全"))
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c + 1).value for c in range(max(col_map.values()) + 1) if True]
        # 读取该行列值
        row_vals = {}
        for field_name, idx in col_map.items():
            v = ws.cell(row=r, column=idx + 1).value
            row_vals[field_name] = "" if v is None else str(v).strip()
        if not any(row_vals.values()):
            continue  # 空行
        qty_raw = row_vals.get("qty", "")
        qty = 0
        try:
            qty = int(float(qty_raw)) if qty_raw not in ("", None) else 0
        except (ValueError, TypeError):
            qty = -1
        sr = SKURow(
            style=row_vals.get("style", ""), sku=row_vals.get("sku", ""),
            sku_name=row_vals.get("sku_name", ""), color=row_vals.get("color", ""),
            size=row_vals.get("size", ""), qty=qty, factory=row_vals.get("factory", ""),
            supplier=row_vals.get("supplier", ""), fabric=row_vals.get("fabric", ""),
            remark=row_vals.get("remark", ""), row_no=r,
        )
        rows.append(sr)
    return rows


# ---------- 面料表解析 ----------

FAB_HEADER_MAP = {
    "款号": "style", "产品名称": "product_name", "品名": "fabric", "纸样名称": "pattern",
    "纸样编号": "pattern", "洗水唛": "wash_label", "每条布出货数": "pcs_per_roll",
    "出货数": "pcs_per_roll", "是否需要压缩袋": "need_bag", "压缩袋": "need_bag",
    "包装袋规格": "bag_spec", "包装规格": "bag_spec",
}


def _parse_wash_components(text: str) -> List[str]:
    """把洗水唛成分字符串解析成「每个成分独占一项」的列表，供生成 Excel 公式拼接用。

    与 _normalize_wash_label 配合：先调用它把同一行的多个成分用 \\n 隔开，再按 \\n 拆分。
    返回的列表形如 ['35% Rayon', '60% Polyester', '5% Spandex']。

    示例:
        "95% Polyester5% Spandex" → ['95% Polyester', '5% Spandex']
        "100% Cotton"             → ['100% Cotton']
        ""                        → []
    """
    if not text:
        return []
    normalized = _normalize_wash_label(text)
    parts = [p.strip() for p in normalized.split("\n")]
    return [p for p in parts if p]


def _normalize_wash_label(text: str) -> str:
    """洗水唛成分自动换行：每个成分通常以"数字%"开头，
    在第二个及之后的百分比前插入换行符，使得每个成分独占一行。

    关键启发式：百分比数字前的字符**不能是数字**（避免把 "95%" 误拆成 "9\n5%"），
    也**不能是换行符**（已有换行的位置不重复加）。

    示例:
        "95% Polyester5% Spandex" → "95% Polyester\\n5% Spandex"
        "95% Rayon\\n5% Spandex"  → "95% Rayon\\n5% Spandex"（已换行，原样保留）
        "100% Cotton"              → "100% Cotton"（单成分，不变）
    """
    if not text:
        return text
    import re as _re
    return _re.sub(r'(?<!^)(?<!\n)(?<!\d)(\d+%)', r'\n\1', text).strip()


def _parse_fabric(ws, errors: List[Tuple[int, str, str, str]]) -> Dict[str, FabricInfo]:
    fabrics: Dict[str, FabricInfo] = {}
    header_row, col_map = None, {}
    for r in range(1, min(ws.max_row, 10) + 1):
        row_vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        if any("款号" in v for v in row_vals) and any("产品名称" in v or "品名" in v or "纸样" in v for v in row_vals):
            header_row = r
            for idx, v in enumerate(row_vals):
                v_low = v.lower().replace(" ", "")
                for key, field_name in FAB_HEADER_MAP.items():
                    if key in v_low:
                        col_map[field_name] = idx
                        break
            break
    if header_row is None:
        errors.append((0, "面料表", "未找到表头行", "面料表需包含 款号/产品名称/品名/纸样名称 等列"))
        return fabrics
    for r in range(header_row + 1, ws.max_row + 1):
        vals = {f: ws.cell(row=r, column=idx + 1).value for f, idx in col_map.items()}
        style = str(vals.get("style") or "").strip()
        if not style:
            continue
        ppr = 40
        try:
            ppr = int(float(vals.get("pcs_per_roll") or 40))
        except (ValueError, TypeError):
            pass
        need_bag_raw = str(vals.get("need_bag") or "").strip()
        need_bag = need_bag_raw in ("是", "Y", "y", "TRUE", "True", "1", "需要")
        wash_text = str(vals.get("wash_label") or "").strip()
        fi = FabricInfo(
            style=style,
            product_name=str(vals.get("product_name") or "").strip(),
            fabric=str(vals.get("fabric") or "").strip(),
            pattern=str(vals.get("pattern") or "").strip(),
            wash_label=_normalize_wash_label(wash_text),
            wash_components=_parse_wash_components(wash_text),
            pcs_per_roll=ppr,
            need_bag=need_bag,
            bag_spec=str(vals.get("bag_spec") or "").strip(),
        )
        fabrics[style] = fi
    return fabrics


# ---------- 配置表解析 ----------

CONFIG_KEYS = {
    "品牌": "brand", "账号代码": "account_code", "账号": "account_code",
    "申请人": "applicant", "联系地址": "office", "办公室": "office",
    "采购仓库": "warehouse", "仓库": "warehouse",
}


def _parse_config(ws) -> Config:
    cfg = Config()
    for r in range(1, min(ws.max_row, 30) + 1):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k is None:
            continue
        k_low = str(k).strip().lower().replace(" ", "")
        for key, field_name in CONFIG_KEYS.items():
            if key in k_low:
                setattr(cfg, field_name, str(v).strip() if v is not None else "")
                break
    return cfg


# ---------- 校验 ----------

def validate(data: InputData):
    errors = data.errors
    rows = data.rows

    # 2. 数据完整性
    for sr in rows:
        if sr.qty < 0:
            errors.append((sr.row_no, "采购数", f"非数字: {sr.qty}", "采购数需为正整数"))
        elif sr.qty == 0:
            errors.append((sr.row_no, "采购数", "采购数为 0", "需 > 0"))
        if sr.size and sr.size not in SIZE_ORDER:
            errors.append((sr.row_no, "尺码", f"未知尺码 {sr.size}", f"应在 {SIZE_ORDER} 内"))
        if not sr.style:
            errors.append((sr.row_no, "款号", "为空", "必填"))
        if not sr.color:
            errors.append((sr.row_no, "颜色", "为空", "必填"))
        if not sr.factory:
            errors.append((sr.row_no, "工厂", "为空", "必填"))
        if not sr.supplier:
            errors.append((sr.row_no, "布行", "为空", "必填"))
        if is_mixed_color(sr.color):
            parts = split_suppliers(sr.supplier)
            if len(parts) != 2:
                errors.append((sr.row_no, "布行", f"混搭色布行需 2 个、逗号/空格/分号分隔, 当前: {sr.supplier!r}",
                               "如: 剑邑布行,宏裕布行 / 剑邑布行 宏裕布行 / 剑邑布行；宏裕布行"))
        elif any(sep in sr.supplier for sep in (",", "，", ";", "；")):
            errors.append((sr.row_no, "布行", f"非混搭色布行不应含分隔符: {sr.supplier!r}", "单色填 1 个布行"))

    # 7. SKU 编码重复
    seen = {}
    for sr in rows:
        if sr.sku:
            if sr.sku in seen:
                errors.append((sr.row_no, "SKU编码", f"与第 {seen[sr.sku]} 行重复", "SKU 编码应唯一"))
            else:
                seen[sr.sku] = sr.row_no

    # 3/4. 同款同色: 工厂一致 & 布行一致（非混搭色）
    by_color: Dict[Tuple[str, str], List[SKURow]] = {}
    for sr in rows:
        key = (sr.style, sr.color)
        by_color.setdefault(key, []).append(sr)
    for (style, color), group in by_color.items():
        if not is_mixed_color(color):
            factories = {sr.factory for sr in group if sr.factory}
            if len(factories) > 1:
                errors.append((group[0].row_no, "工厂",
                               f"同款同色 {style} / {color} 出现多个工厂: {sorted(factories)}",
                               "一般以颜色为最小单位下单，疑似复制错误"))
            suppliers = {sr.supplier for sr in group if sr.supplier}
            if len(suppliers) > 1:
                errors.append((group[0].row_no, "布行",
                               f"同款同色 {style} / {color} 出现多个布行: {sorted(suppliers)}",
                               "疑似复制错误"))

    # 6. 面料表/配置表缺失
    styles = {sr.style for sr in rows}
    for st in styles:
        if st not in data.fabrics:
            errors.append((0, "面料表", f"款号 {st} 在面料表缺失", "面料表需包含每个款的面料信息"))
        else:
            fi = data.fabrics[st]
            if not fi.product_name:
                errors.append((0, "面料表", f"款号 {st} 缺产品名称", "必填"))
            if not fi.pattern:
                errors.append((0, "面料表", f"款号 {st} 缺纸样名称", "必填"))
    if not data.config.brand:
        errors.append((0, "配置表", "品牌为空", "必填"))
    if not data.config.account_code:
        errors.append((0, "配置表", "账号代码为空", "必填"))


# ---------- 图片 sheet 解析 ----------

def _load_cellimages(path: str) -> Dict[str, bytes]:
    """从 xlsx 底层 cellimages.xml 提取嵌入单元格图片 (DISPIMG) 的二进制
    兼容新版 WPS/Excel 用「置于单元格」插入的图片"""
    import zipfile
    out: Dict[str, bytes] = {}
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            if "xl/cellimages.xml" not in names or "xl/_rels/cellimages.xml.rels" not in names:
                return out
            cx = z.read("xl/cellimages.xml").decode("utf-8", "ignore")
            rels = z.read("xl/_rels/cellimages.xml.rels").decode("utf-8", "ignore")
            import re as _re
            rid2media = dict(_re.findall(r'Id="(rId\d+)"[^>]*Target="(media/[^"]+)"', rels))
            for m in _re.finditer(r"<etc:cellImage>(.*?)</etc:cellImage>", cx, _re.S):
                body = m.group(1)
                idm = _re.search(r'name="(ID_[0-9A-F]+)"', body)
                rid = _re.search(r'r:embed="(rId\d+)"', body)
                if idm and rid and rid.group(1) in rid2media:
                    out[idm.group(1)] = z.read("xl/" + rid2media[rid.group(1)])
    except Exception:
        pass
    return out


def _parse_images(ws, path: str = None) -> Dict[str, bytes]:
    """读取「图片」sheet 的图片 -> {位置名: bytes}。约定:
    A2~A5 为标签, 图片左上角锚定 B2~B5。
    兼容两种插入方式:
      ① 普通浮动图片 (openpyxl 直接读 _images)
      ② DISPIMG 嵌入单元格 (新版 WPS「置于单元格」,从 cellimages.xml 解析)"""
    import re as _re
    out: Dict[str, bytes] = {}
    if ws is None:
        return out
    # ① 普通浮动图片
    for img in getattr(ws, "_images", []):
        try:
            row1 = img.anchor._from.row + 1
        except (AttributeError, TypeError):
            continue
        label = IMAGE_LABELS.get(row1)
        if label:
            try:
                out[label] = img._data()
            except Exception:
                continue
    # ② DISPIMG 嵌入单元格图片（兜底）—— 需要 data_only=False 看公式字符串
    if path:
        cellimg = _load_cellimages(path)
        if cellimg:
            try:
                wb_f = openpyxl.load_workbook(path, data_only=False)
                ws_f = None
                # 找「图片」sheet 对应的 ws
                if ws.title in wb_f.sheetnames:
                    ws_f = wb_f[ws.title]
                if ws_f is not None:
                    for row in ws_f.iter_rows():
                        for cell in row:
                            v = cell.value
                            if not isinstance(v, str) or "DISPIMG" not in v.upper():
                                continue
                            m = _re.search(r'DISPIMG\("?(ID_[0-9A-Fa-f]+)"?', v, _re.I)
                            if m and m.group(1) in cellimg:
                                label = IMAGE_LABELS.get(cell.row)
                                if label and label not in out:
                                    out[label] = cellimg[m.group(1)]
            except Exception:
                pass
    return out


# ---------- 入口 ----------

def parse_input(path: str) -> InputData:
    wb = openpyxl.load_workbook(path, data_only=True)
    data = InputData(input_filename=path.split("\\")[-1].split("/")[-1])

    ws_main = _find_sheet(wb, ["主表", "SKU", "采购"], 0)
    ws_fab = _find_sheet(wb, ["面料", "纸样", "布料"], 1)
    ws_cfg = _find_sheet(wb, ["配置", "品牌"], 2)
    ws_img = _find_sheet(wb, ["图片", "image"], 3)

    if ws_main is None:
        data.errors.append((0, "主表", "未找到主表", "需要包含款号/SKU的主表"))
    else:
        data.rows = _parse_main(ws_main, data.errors)
    if ws_fab is None:
        data.errors.append((0, "面料表", "未找到面料表", "需要面料信息表"))
    else:
        data.fabrics = _parse_fabric(ws_fab, data.errors)
    if ws_cfg is None:
        data.errors.append((0, "配置表", "未找到配置表", "需要品牌/账号代码/申请人/联系地址"))
    else:
        data.config = _parse_config(ws_cfg)
    if ws_img is not None:
        data.images = _parse_images(ws_img, path)

    validate(data)
    return data
