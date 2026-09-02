# -*- coding: utf-8 -*-
"""generator.py — 输出文件生成

以模板为底版:
  - 订单 sheet:  复制 THD 模板「订单」sheet（保留样式/图片/底部固定内容），重写头部+主体
  - 布料申购单:  复制 PPY 旧模板「布料申购单」sheet，每个布行一张，删除右块，填数据
  - 汇总明细表:  独立生成 文件汇总 + 布行明细
"""
import os
import re
import shutil
import zipfile
from copy import copy
from datetime import datetime, date as date_cls
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.drawing.image import Image as XLImage

from .splitter import OrderUnit, SplitResult
from .parser import InputData

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
THD_TEMPLATE = os.path.join(BASE_DIR, "templates", "thd_template_clean.xlsx")  # 净化模板: 图片已内置
PPY_TEMPLATE = os.path.join(BASE_DIR, "templates", "ppy_old_template.xlsx")
IMAGES_DIR = os.path.join(BASE_DIR, "templates", "images")

THIN = Side(style="thin", color="000000")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 模板内嵌图片（DISPIMG）→ 普通图片插回：位置名 / 默认图片ID / 目标紧凑尺寸(px)
# 比 THD 模板原 DISPIMG ext 更小（模板原图过大导致超出视觉区域）
DISPIMG_DEFAULTS = {
    "包装袋图片": ("ID_BE270A7E788F4E49844F73A78120EB23", 350, 230),
    "主唛": ("ID_6F3DBA6A97AB49F3AC43052A1ECB3BFC", 320, 240),
    "吊牌": ("ID_9C0A6E5FB0764A84BB59F70FBFA53CD6", 260, 110),
    "温馨提示卡片": ("ID_2A34054940634D1BAF0639E8B12A910E", 300, 250),
}


def _locate_image_anchors(ws) -> Dict[str, Tuple[int, int]]:
    """按标签行动态定位图片锚点（颜色行多时固定内容下移也能找到）
    规则: '包装袋图片'(A列标签) -> 图在其下一行 A 列; '主唛'(A列) -> 下一行 A 列;
          '吊牌'(D列标签) -> 下一行 D 列; '温馨提示'(B列) -> 本行 C 列"""
    anchors: Dict[str, Tuple[int, int]] = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        d = ws.cell(row=r, column=4).value
        b = ws.cell(row=r, column=2).value
        if isinstance(a, str) and "包装袋图片" in a:
            anchors["包装袋图片"] = (r + 1, 1)
        if isinstance(a, str) and "主唛" in a and "吊牌" not in a:
            anchors["主唛"] = (r + 1, 1)
        if isinstance(d, str) and "吊牌" in d:
            anchors["吊牌"] = (r + 1, 4)
        if b is not None and "温馨提示" in str(b):
            anchors["温馨提示卡片"] = (r, 3)
    return anchors


def _shift_images_down(ws, start_1based: int, extra: int):
    """插行后把锚点位于插行位置及其下方的图片整体下移 extra 行（0-based 处理）"""
    if extra <= 0:
        return
    for img in ws._images:
        try:
            r0 = img.anchor._from.row
        except (AttributeError, TypeError):
            continue
        if r0 >= start_1based - 1:
            img.anchor._from.row = r0 + extra


def _shift_merges_down(ws, start_1based: int, extra: int):
    """插行后把位于插行位置及其下方的合并区域整体下移 extra 行。
    ⚠️ 关键修复：openpyxl 的 `unmerge_cells` 会删除 _cells dict 中的 MergedCell 引用，
       但延伸区域可能包含物理值（insert_rows 把值移到新位置但延伸 MergedCell 已被新位置覆盖）。
       更糟糕的是，unmerge_cells 会**误删**新位置上不在合并范围内的普通 Cell（因为延伸范围可能
       包括了 insert_rows 后物理移动到那里的普通 Cell，例如 G46=O51 合并延伸 cell G47:G50 等，
       但 G46 是 D46:H46 的延伸 cell 也包含 G46，会被误删）。

       **采用 in-place shift**：直接修改 MergedCellRange 对象的坐标，值已经物理移动到新位置，
       这样完全不破坏 _cells 字典、不清空值、不需要重新 merge。

    同时**手动移动行高**——openpyxl 的 row_dimensions dict key 不会随 insert_rows 移动，
       必须先清空原位置行高，再设到新位置。"""
    if extra <= 0:
        return
    # 1. 收集所有 >= start_1based 的合并区域
    affected = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= start_1based:
            affected.append(mr)
    # 2. 收集所有 >= start_1based 的行高（按行）
    row_heights = {}
    for r in range(start_1based, ws.max_row + 1):
        h = ws.row_dimensions[r].height
        if h is not None:
            row_heights[r] = h
    # 3. ⚠️ In-place shift（最干净的方式）：直接修改 MergedCellRange 坐标到新位置
    #    值已经在新位置（insert_rows 物理移动），不需要清空/恢复。
    for mr in affected:
        try:
            mr.shift(row_shift=extra)
        except Exception:
            pass
    # 4. 清空原位置的行高（不会动到值）
    for r in range(start_1based, ws.max_row + 1):
        ws.row_dimensions[r].height = None
    # 5. 行高按"按行"批量重新设置到下移后的位置——精细控制每一行
    for old_r, h in row_heights.items():
        ws.row_dimensions[old_r + extra].height = h


def _shift_merges_up(ws, start_1based: int, amount: int):
    """删除行后把位于删除位置及其下方的合并区域整体上移 amount 行。
    与 _shift_merges_down 相对：使用 mr.shift(row_shift=-amount)。
    注：openpyxl 的 delete_rows 不会自动处理 merged_cells.ranges，需手动 shift。
    行高的迁移由调用方在 delete_rows 之前收集、之后迁移（见 _delete_empty_rows_below）。"""
    if amount <= 0:
        return
    affected = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= start_1based:
            affected.append(mr)
    for mr in affected:
        try:
            mr.shift(row_shift=-amount)
        except Exception:
            pass


def _shift_images_up(ws, start_1based: int, amount: int):
    """删除行后把锚点位于删除位置及其下方的图片整体上移 amount 行（0-based 处理）"""
    if amount <= 0:
        return
    for img in ws._images:
        try:
            r0 = img.anchor._from.row
        except (AttributeError, TypeError):
            continue
        if r0 >= start_1based - 1:
            new_r = r0 - amount
            if new_r >= 0:
                img.anchor._from.row = new_r


def _delete_empty_rows_below(ws, total_row: int, fixed_top: int):
    """删除总计行与固定内容（包装袋图片）之间的空白行。
    调用 ws.delete_rows 删除行后, 手动上移合并区域和图片锚点（openpyxl 的
    delete_rows 不会自动更新 _images 锚点和 merged_cells.ranges）。
    ⚠️ 行高必须在 delete_rows **之前**收集：openpyxl 的 delete_rows 不清理
    row_dimensions dict 中的旧 key，且会把 ws.max_row 变小——若删除后再收集，
    原 max_row 以下的残留行高（如模板 R49=80/R53=31.5/R60=78/R61=69）会因
    超过新 max_row 而漏收，导致固定区行高全部丢失。
    返回新的 fixed_top（已上移 amount 行）。"""
    empty_count = fixed_top - total_row - 1
    if empty_count <= 0:
        return fixed_top
    delete_start = total_row + 1   # 1-based

    # 0. ⚠️ 删除前收集行高（只保留删除范围**以下**的固定区行高）
    keep_heights = {}
    for r in range(delete_start + empty_count, ws.max_row + 1):
        h = ws.row_dimensions[r].height
        if h is not None:
            keep_heights[r] = h

    # 1. 删除行（openpyxl 只移动单元格值）
    ws.delete_rows(delete_start, empty_count)

    # 2. ⚠️ 迁移行高：delete_rows 不清理 row_dimensions——先清空全部残留
    #    （>= delete_start 的旧 key，包括被删行范围的 25.0 数据区行高），再写新位置
    for r in list(ws.row_dimensions.keys()):
        if r >= delete_start:
            ws.row_dimensions[r].height = None
    for old_r, h in keep_heights.items():
        ws.row_dimensions[old_r - empty_count].height = h

    # 3. 兜底：上移图片锚点（openpyxl delete_rows 不会更新 _images）
    _shift_images_up(ws, delete_start, empty_count)
    # 4. 兜底：上移合并区域（delete_rows 不处理 merged_cells.ranges）
    _shift_merges_up(ws, delete_start, empty_count)
    # 5. 返回新的 fixed_top
    return fixed_top - empty_count


def _capture_row_style(ws, row: int, col_start: int = 1, col_end: int = 15):
    """捕获指定行的单元格样式（font/border/fill/alignment/number_format）和行高。
    用于插入行后把样板行的样式复制到新行（保持颜色行/汇总行带边框/底纹）。"""
    from copy import copy
    captured = []
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row, c)
        captured.append({
            "font": copy(cell.font) if cell.font else None,
            "border": copy(cell.border) if cell.border else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        })
    row_height = ws.row_dimensions[row].height
    return {"cells": captured, "row_height": row_height, "col_start": col_start, "col_end": col_end}


def _apply_row_style(ws, style: dict, row: int):
    """把捕获的样式应用到指定行（行高 + 单元格样式）"""
    from openpyxl.cell.cell import MergedCell
    cells = style["cells"]
    for i, c in enumerate(range(style["col_start"], style["col_end"] + 1)):
        cell = ws.cell(row, c)
        if isinstance(cell, MergedCell):
            continue
        s = cells[i]
        if s["font"]: cell.font = s["font"]
        if s["border"]: cell.border = s["border"]
        if s["fill"]: cell.fill = s["fill"]
        if s["alignment"]: cell.alignment = s["alignment"]
        cell.number_format = s["number_format"]
    if style["row_height"]:
        ws.row_dimensions[row].height = style["row_height"]


def _cell_area_size_px(ws, row: int, col: int) -> Tuple[float, float]:
    """计算锚点 (row, col) 所在单元格/合并区域的实际像素尺寸 (宽, 高)。
    ⚠️ 图片要"规整放进单元格"，必须按目标区域的实际列宽×行高换算像素，
       不能用硬编码尺寸（否则图片超出边框）。"""
    from openpyxl.utils import get_column_letter
    # 1. 找包含该格的合并区域（图片放合并区域内更规整）
    area = None
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            area = mr
            break
    r1, r2 = (area.min_row, area.max_row) if area else (row, row)
    c1, c2 = (area.min_col, area.max_col) if area else (col, col)
    # 2. 总列宽（Excel 字符单位）→ 像素：默认 11pt 字体下 1 字符宽 ≈ 7px
    total_w = 0.0
    for c in range(c1, c2 + 1):
        w = ws.column_dimensions[get_column_letter(c)].width
        total_w += w if w else 8.43   # 默认列宽 8.43 字符
    width_px = total_w * 7
    # 3. 总行高（pt）→ 像素：1pt = 96/72 px（96 DPI 屏幕）
    total_h = 0.0
    for r in range(r1, r2 + 1):
        h = ws.row_dimensions[r].height
        total_h += h if h else 15.0   # 默认行高 15pt
    height_px = total_h * 96 / 72
    return width_px, height_px


def _build_image_anchor(ws, row: int, col: int, image_w: int, image_h: int):
    """为图片构造锚点：等比缩放到（合并）区域 + 居中放置。
    采用 OneCellAnchor + colOff/rowOff 偏移（openpyxl 最标准的图片放置方式，
    跨 Excel/WPS 渲染一致）— TwoCellAnchor 实际渲染时图片会偏左上不居中。
    返回 (anchor, new_w, new_h) 三元组。"""
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    # 1. 找合并区域
    area = None
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            area = mr
            break
    if area:
        r1, r2 = area.min_row, area.max_row
        c1, c2 = area.min_col, area.max_col
    else:
        r1 = r2 = row
        c1 = c2 = col
    # 2. 计算区域总像素宽高
    area_w = sum((ws.column_dimensions[get_column_letter(c)].width or 8.43) * 7
                 for c in range(c1, c2 + 1))
    area_h = sum((ws.row_dimensions[r].height or 15.0) * 96 / 72
                 for r in range(r1, r2 + 1))
    # 3. 等比缩放（留 10% 内边距，宽限 1 像素避免溢出）
    scale = min(area_w * 0.90 / image_w, area_h * 0.90 / image_h)
    iw = max(int(image_w * scale), 1)
    ih = max(int(image_h * scale), 1)
    # 4. 居中偏移（按区域左上角，colOff/rowOff 是 EMU）
    off_x = pixels_to_EMU(int((area_w - iw) / 2))
    off_y = pixels_to_EMU(int((area_h - ih) / 2))
    # 5. OneCellAnchor：_from.colOff/rowOff 是从该单元格的偏移；ext 是图片总尺寸
    anchor = OneCellAnchor(
        _from=AnchorMarker(col=c1 - 1, colOff=off_x, row=r1 - 1, rowOff=off_y),
        ext=XDRPositiveSize2D(pixels_to_EMU(iw), pixels_to_EMU(ih))
    )
    return anchor, iw, ih


def _replace_cell_images(ws, unit_images: Dict[str, bytes] = None):
    """仅当输入文件/网页提供了替换图时，替换模板中对应位置的图片。
    模板里的默认图保留不动（方案2: 图已内置在净化模板中）。
    ⚠️ 采用 TwoCellAnchor + 等比居中：图片嵌入（合并）单元格区域中央，
       缩放比例最优、不变形、不超出边框、跨 Excel/WPS 一致。"""
    import io
    unit_images = unit_images or {}
    if not unit_images:
        return
    anchors = _locate_image_anchors(ws)
    # 1. 删除模板中"有替换图"位置的旧图
    for img in list(ws._images):
        try:
            r0, c0 = img.anchor._from.row, img.anchor._from.col
        except (AttributeError, TypeError):
            continue
        for label, (row, col) in anchors.items():
            if r0 == row - 1 and c0 == col - 1 and label in unit_images:
                ws._images.remove(img)
                break
    # 2. 插入新图（TwoCellAnchor + 等比居中嵌入区域）
    for label, (row, col) in anchors.items():
        if label not in unit_images:
            continue
        nimg = XLImage(io.BytesIO(unit_images[label]))
        iw, ih = nimg.width, nimg.height
        if iw and ih:
            anchor, new_w, new_h = _build_image_anchor(ws, row, col, iw, ih)
            nimg.width, nimg.height = new_w, new_h
            nimg.anchor = anchor
        ws.add_image(nimg)


def _clear_data_values(ws, fixed_top):
    """只清空数据区（R8~fixed_top-1）的单元格值，**完全保留样式/合并/行高/列宽/边框**。
    用户要求：格式以样板为准，只管写数据。"""
    from openpyxl.cell.cell import MergedCell
    for r in range(8, fixed_top):
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _find_fixed_top(ws, start_row=8):
    """动态定位固定内容起点：找到「包装袋图片」标签行（样板在 R36 或 R24）。
    数据区 = R8 ~ fixed_top-1。"""
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "包装袋" in v:
            return r
    return ws.max_row + 1  # 找不到则用最大行


def _to_date_value(value):
    """把日期值统一转成 date 对象（用于写入订单「日期」/申购单「申请日期」，显示 YYYY/M/D）。

    兼容输入:
      - 8 位串   '20260902'        -> date(2026, 9, 2)
      - 带分隔符 '2026/09/02' / '2026-09-02' / '2026.09.02'（ui 手动编辑常见）-> date(2026, 9, 2)
      - date / datetime 对象直接返回
    解析失败（含空值）返回原值，不抛异常 —— 避免单个日期异常导致整单生成崩。
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date_cls):
        return value
    s = str(value or "").strip()
    if not s:
        return value
    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%Y%m%d").date()
        except ValueError:
            return value
    return value


def _write_date_cell(ws, row: int, col: int, value):
    """写日期值到单元格，并强制 yyyy/m/d 显示格式（真正的日期值，可排序/筛选）。
    月份/日期为单个数字时不补 0（如 2026/9/2），两位数照常显示（如 2026/10/15）。
    传入非日期值（如日期串解析失败）则原样写入，不设格式。"""
    cell = ws.cell(row=row, column=col)
    dv = _to_date_value(value)
    cell.value = dv
    if isinstance(dv, (datetime, date_cls)):
        cell.number_format = "yyyy/m/d"
    return cell


# ---------- 订单 sheet ----------

def _build_order_sheet(wb, unit: OrderUnit):
    """以样板为底版，**只写数据，完全不动格式**（列宽/行高/合并/边框/填充全部保留样板原样）。
    数据区布局（样板固定）：C~K=9尺码(S~6XL), L=比例, M=总计, N=布料条数, O=备注
    """
    ws = wb["订单"]

    # 0. 动态定位固定内容起点（模板样板可能是 1色版 R24 或 26色版 R36）
    fixed_top = _find_fixed_top(ws)
    if fixed_top < 10:
        fixed_top = 24   # 兜底：找不到「包装袋图片」标签时用模板默认

    # 1. 只清空数据区 R8~fixed_top-1 的值（保留全部格式）
    _clear_data_values(ws, fixed_top)

    # 2. 头部字段（只写值，样板已有标签/合并/样式）
    ws["B2"] = unit.style          # 款号
    ws["H2"] = unit.product_name   # 名称
    ws["B3"] = unit.order_no       # 订单号
    ws["H3"] = unit.config_brand   # 品牌
    ws["B4"] = unit.pattern        # 做货纸样编号
    ws["B5"] = unit.factory        # 加工厂
    _write_date_cell(ws, 5, 8, unit.date_str)  # H5 日期，显示 YYYY/M/D（如 2026/9/2，单数字月/日不补 0）

    # 3. 表头：**不重写 A6/B6/L6/M6/N6/O6**（模板已有正确值，含 A6 斜线表头的空格布局）
    #    只写尺码名 C~K（值一致，不碰样式）
    sizes = unit.sizes   # 固定 ['S','M','L','XL','2XL','3XL','4XL','5XL','6XL']
    for i, s in enumerate(sizes):
        ws.cell(row=6, column=3 + i).value = s
    # ⚠️ 强制设置 R6 行高：容纳 24pt 字号两行（"颜色"+"尺码"），模板默认 25 太小
    ws.row_dimensions[6].height = 50.0

    # 4. 数据行（R8 起）+ 汇总行（隔一行）
    start = 8
    n = len(unit.color_rows)
    # fixed_top 已在步骤0动态检测（样板 26色版在 R36；1色版在 R24）
    avail = fixed_top - start
    need_rows = n + 2   # 数据行 + 空行 + 汇总行
    extra = need_rows - avail if need_rows > avail else 0

    # ⚠️ 关键顺序：先 insert → _shift_merges_down（内部 unmerge+清空行高+重新 merge+恢复行高）→ _apply_row_style
    #    注意：**不能预先 unmerge**！否则 _shift_merges_down 内部 affected 收集为 0，导致合并区域全丢
    if extra > 0:
        # 0. 保存原始 fixed_top（_shift_images_down 仍需旧值定位图片下移起点）
        fixed_top_original = fixed_top
        # 1. 插入行
        ws.insert_rows(fixed_top, extra)
        # 2. 移动固定内容的合并区域 + 行高（必须在 _apply_row_style 之前！否则新空行会被 R8 样式覆盖）
        _shift_merges_down(ws, fixed_top, extra)
        # 3. 捕获 R8 样式 + 给新插入行应用 R8 样式
        template_style = _capture_row_style(ws, 8, 1, 15)
        for r in range(fixed_top, fixed_top + extra):
            _apply_row_style(ws, template_style, r)
        # 4.5 移动图片（用原始 fixed_top，因为 openpyxl 的 _images 锚点未随 insert_rows 更新）
        _shift_images_down(ws, fixed_top_original, extra)
        # ⚠️ 5. 更新 fixed_top：包装袋图片等固定内容已下移 extra 行
        fixed_top += extra

    # 6. 写颜色行（值）
    for idx, cr in enumerate(unit.color_rows):
        r = start + idx
        ws.cell(row=r, column=1, value=cr.color)
        ws.cell(row=r, column=2, value=cr.bag_spec)
        for i, s in enumerate(sizes):
            ws.cell(row=r, column=3 + i, value=cr.qty_by_size.get(s) or None)
        ws.cell(row=r, column=3 + len(sizes), value=cr.ratio)      # 比例
        ws.cell(row=r, column=4 + len(sizes), value=cr.total)      # 总计
        ws.cell(row=r, column=5 + len(sizes), value=cr.rolls)      # 布料条数
        ws.cell(row=r, column=6 + len(sizes), value=cr.remark)     # 备注

    # 7. 汇总行（值）
    blank_row = start + n
    for c in range(1, 7 + len(sizes)):
        ws.cell(row=blank_row, column=c).value = None
    total_row = start + n + 1
    ws.cell(row=total_row, column=1, value="总计")
    ws.cell(row=total_row, column=4 + len(sizes), value=unit.total_qty)
    ws.cell(row=total_row, column=5 + len(sizes), value=sum(cr.rolls for cr in unit.color_rows))

    # 8. 写完颜色行/汇总行后，把固定内容（合并+行高+图片）已通过 _shift_merges_down 处理
    # 4.5 已调用 _shift_merges_down（在 extra>0 时）；以下手动的 remerge 逻辑已废弃

    # 9. 洗水唛内容（纯值写入，动态定位「洗水唛内容」标签行，值写在下一行 A 列）
    #     原模板是公式: =VLOOKUP(B2,'THD款式图（请及时更新）'!A:C,3,FALSE)&"
    #                    +常规洗涤方式
    #                    MADE IN CHINA"
    #     用户要求: 不需要公式，直接按 Python 计算结果写值即可——成分取输入文件的
    #              洗水唛字段（每个成分独占一行），成分与文字常量之间加一行 "+"，
    #              最后固定追加 "常规洗涤方式" 和 "MADE IN CHINA"（v26）。
    #     成分缺失（输入文件未填）→ 不写成分和 "+" 行，只保留两行常量（网页会提示补成分）。
    for r in range(fixed_top, min(ws.max_row, 90) + 1):
        if ws.cell(row=r, column=1).value and "洗水唛" in str(ws.cell(row=r, column=1).value):
            cell = ws.cell(row=r + 1, column=1)
            # 优先用成分列表（保持输入文件原顺序）；退路用 wash_label 按 \n 拆分
            components = list(unit.wash_components or [])
            if not components and unit.wash_label:
                components = [p.strip() for p in unit.wash_label.split("\n") if p.strip()]
            if components:
                # 输出格式（图2）：成分行... / "+" / 常规洗涤方式 / MADE IN CHINA
                cell.value = "\n".join(components + ["+", "常规洗涤方式", "MADE IN CHINA"])
            else:
                # 成分为空：不写成分行和 "+"，只保留固定常量
                cell.value = "\n".join(["常规洗涤方式", "MADE IN CHINA"])
            # ⚠️ 强制 wrap_text=True: 模板默认 wrap=None 时多行内容会撑高, 这里显式确保换行
            from openpyxl.styles import Alignment
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                      vertical=cell.alignment.vertical,
                                      wrap_text=True)
            break

    # 10. 外包装袋贴纸内容（公式实现，动态定位「外包装袋贴纸内容」标签）
    #     公式: ="款号-颜色-尺码"&CHAR(10)&"名称"&CHAR(10)&"例如:"&CHAR(10)&B2&"-黑色-S"&CHAR(10)&H2
    #     B2 = 款号, H2 = 名称（实际值由 Excel/WPS 求值得出: 5 行拼接）
    PACKING_FORMULA = '="款号-颜色-尺码"&CHAR(10)&"名称"&CHAR(10)&"例如:"&CHAR(10)&B2&"-黑色-S"&CHAR(10)&H2'
    for r in range(fixed_top, min(ws.max_row, 90) + 1):
        for c in range(1, 16):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "外包装袋贴纸" in v:
                cell = ws.cell(row=r + 1, column=c)
                cell.value = PACKING_FORMULA
                from openpyxl.styles import Alignment
                cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                          vertical=cell.alignment.vertical,
                                          wrap_text=True)
                break
        else:
            continue
        break

    # 11. ⚠️ 删除总计行与包装袋图片之间的空白行（v24 修复）
    #     注意：必须在所有固定内容写入完成后调用，确保图片锚点和合并区域都已稳定
    fixed_top = _delete_empty_rows_below(ws, total_row, fixed_top)

    # 6. 仅当输入文件/网页提供了替换图时，替换模板中对应位置的图片（模板默认图保留）
    _replace_cell_images(ws, getattr(unit, "images", None) or {})
    return ws


# ---------- 布料申购单 sheet ----------

def _build_supplier_sheet(wb, unit: OrderUnit, group):
    """从 PPY 模板布料申购单复制样式并生成一个布行申购单 sheet"""
    src = openpyxl.load_workbook(PPY_TEMPLATE, data_only=False)
    src_ws = src["布料申购单"]
    new_ws = wb.create_sheet(title=f"布料申购单-{group.supplier}")

    # 复制列宽
    for col, dim in src_ws.column_dimensions.items():
        new_ws.column_dimensions[col].width = dim.width
    # 复制行高 + 单元格值/样式（仅左块 A~D 列 + 表头，右块 H~K 不复制）
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=4):
        for c in row:
            if c.value is not None:
                new_ws.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                new_ws.cell(row=c.row, column=c.column).font = copy(c.font)
                new_ws.cell(row=c.row, column=c.column).border = copy(c.border)
                new_ws.cell(row=c.row, column=c.column).fill = copy(c.fill)
                new_ws.cell(row=c.row, column=c.column).alignment = copy(c.alignment)
            new_ws.row_dimensions[c.row].height = src_ws.row_dimensions[c.row].height
    # 复制左块合并单元格（A1:D1, A2:D2... 检查）
    for mr in src_ws.merged_cells.ranges:
        if mr.min_col <= 4:   # 只复制左块
            new_ws.merge_cells(str(mr))

    # 填头部
    new_ws["B2"] = unit.order_no
    new_ws["D2"] = unit.style
    new_ws["B3"] = ""            # 货号 留空
    new_ws["D3"] = ""            # 重量 留空
    new_ws["B4"] = unit.fabric   # 品名
    new_ws["D4"] = unit.factory  # 代加工工厂
    new_ws["B8"] = unit.config_applicant
    _write_date_cell(new_ws, 8, 4, unit.date_str)   # D8 申请日期，显示 YYYY/M/D

    # 数据行（R6 起，模板 R6 是第 1 条数据）
    data_start = 6
    rows = group.rows
    need = len(rows)
    # 模板数据行从 R6 到 R6（1 行），合计在 R7，申请人在 R8
    # 插入额外行使数据区容纳所有颜色
    if need > 1:
        new_ws.insert_rows(data_start + 1, need - 1)
    # 以 R6 为样式基准，复制到所有数据行（修复 insert_rows 后样式不一致）
    for i, sr in enumerate(rows):
        r = data_start + i
        new_ws.cell(row=r, column=1, value=sr.color)
        new_ws.cell(row=r, column=2, value=sr.rolls)
        new_ws.cell(row=r, column=3, value=sr.supplier)
        new_ws.cell(row=r, column=4, value=unit.config_account_code)
        # 复制 R6 的样式（边框/字体/对齐/填充）到 R(i)，确保与 R6 一致
        for c in range(1, 5):
            src_c = new_ws.cell(row=data_start, column=c)
            dst_c = new_ws.cell(row=r, column=c)
            if src_c.has_style:
                dst_c.font = copy(src_c.font)
                dst_c.alignment = copy(src_c.alignment)
                dst_c.fill = copy(src_c.fill)
                dst_c.border = copy(src_c.border)
        new_ws.row_dimensions[r].height = new_ws.row_dimensions[data_start].height
    # 合计行
    total_row = data_start + need
    new_ws.cell(row=total_row, column=1, value="合计")
    new_ws.cell(row=total_row, column=2, value=group.total_rolls)
    # 申请人行（若插行后申请人行被下移，重定位）
    new_ws.cell(row=total_row + 1, column=1, value="申请人")
    new_ws.cell(row=total_row + 1, column=2, value=unit.config_applicant)
    new_ws.cell(row=total_row + 1, column=3, value="申请日期")
    _write_date_cell(new_ws, total_row + 1, 4, unit.date_str)   # 申请日期，显示 YYYY/M/D
    # 合计行 / 申请人行 应用与数据行一致的样式（字体/对齐/边框/行高）
    for rr in (total_row, total_row + 1):
        for c in range(1, 5):
            src_c = new_ws.cell(row=data_start, column=c)
            dst_c = new_ws.cell(row=rr, column=c)
            if src_c.has_style:
                dst_c.font = copy(src_c.font)
                dst_c.alignment = copy(src_c.alignment)
                dst_c.fill = copy(src_c.fill)
                dst_c.border = copy(src_c.border)
        new_ws.row_dimensions[rr].height = new_ws.row_dimensions[data_start].height
    return new_ws


# ---------- 汇总明细表 ----------

def build_summary(split_result: SplitResult, date_str: str, out_dir: str) -> str:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "文件汇总"
    headers1 = ["订单号", "款号", "产品名称", "类型", "工厂", "涉及布行", "颜色数", "总件数", "布料总条数", "生成文件名"]
    ws1.append(headers1)
    for u in split_result.units:
        suppliers = "、".join(g.supplier for g in u.supplier_groups)
        ws1.append([u.order_no, u.style, u.product_name, u.type_str, u.factory,
                    suppliers, len(u.color_rows), u.total_qty,
                    sum(g.total_rolls for g in u.supplier_groups), u.file_name])

    ws2 = wb.create_sheet("布行明细")
    headers2 = ["订单号", "布行", "品名", "颜色数", "条数合计"]
    ws2.append(headers2)
    for u in split_result.units:
        for g in u.supplier_groups:
            ws2.append([u.order_no, g.supplier, g.fabric, len(g.rows), g.total_rolls])

    for ws, ncol in ((ws1, len(headers1)), (ws2, len(headers2))):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.border = BOX
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = BOX
        ws.freeze_panes = "A2"

    path = os.path.join(out_dir, f"{date_str}_采购单汇总.xlsx")
    wb.save(path)
    return path


# ---------- 入口 ----------

def generate_all(data: InputData, split_result: SplitResult, out_dir: str) -> List[str]:
    os.makedirs(out_dir, exist_ok=True)
    date_str = split_result.units[0].order_no[3:11] if split_result.units else data.input_filename
    paths = []
    for unit in split_result.units:
        unit.config_brand = data.config.brand
        unit.config_applicant = data.config.applicant
        unit.config_account_code = data.config.account_code
        unit.config_office = data.config.office
        unit.date_str = date_str
        unit.images = data.images   # 输入文件「图片」sheet 的图（可能为空 → 用默认）
        # 复制 THD 模板
        wb = openpyxl.load_workbook(THD_TEMPLATE, data_only=False)
        # 删除除「订单」外所有 sheet
        for ws in list(wb.worksheets):
            if ws.title != "订单":
                wb.remove(ws)
        _build_order_sheet(wb, unit)
        for g in unit.supplier_groups:
            _build_supplier_sheet(wb, unit, g)
        path = os.path.join(out_dir, unit.file_name)
        wb.save(path)
        paths.append(path)

    # 汇总表
    build_summary(split_result, date_str, out_dir)
    return paths


def make_zip(out_dir: str, zip_path: str, date_str: str) -> str:
    """打包 out_dir 下所有 xlsx 为 zip"""
    files = [f for f in os.listdir(out_dir) if f.endswith(".xlsx")]
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in files:
            zf.write(os.path.join(out_dir, f), arcname=f)
    return zip_path
