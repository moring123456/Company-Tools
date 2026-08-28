# -*- coding: utf-8 -*-
"""生成净化模板: 把订单 sheet 的 DISPIMG 单元格图片替换为普通浮动图片
保存为 templates/thd_template_clean.xlsx，generator 以它为底版。
之后用户可直接在 WPS/Excel 打开此模板修改图片位置/内容"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from core.generator import _locate_image_anchors, DISPIMG_DEFAULTS, IMAGES_DIR

SRC = os.path.join("templates", "thd_template.xlsx")
DST = os.path.join("templates", "thd_template_clean.xlsx")


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)

    # 0. 删除除「订单」外的所有 sheet（防止 SUMIFS 等公式跨表引用报错）
    for sn in list(wb.sheetnames):
        if sn != "订单":
            del wb[sn]

    ws = wb["订单"]

    # 1. 清除订单 sheet 全部公式（=开头）和 DISPIMG 字符串（保留样式/合并/图片）
    cleared_formula = 0
    cleared_dispimg = 0
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str):
                if v.startswith("="):
                    c.value = None
                    cleared_formula += 1
                elif "DISPIMG" in v.upper():
                    c.value = None
                    cleared_dispimg += 1

    # 2. 按标签定位锚点，插入普通浮动图片（紧凑尺寸）
    anchors = _locate_image_anchors(ws)
    inserted_imgs = 0
    for label, (img_id, w, h) in DISPIMG_DEFAULTS.items():
        anchor = anchors.get(label)
        if not anchor:
            print(f"  ⚠️ 未找到 {label} 标签位置，跳过")
            continue
        row, col = anchor
        p = os.path.join(IMAGES_DIR, f"{img_id}.png")
        if not os.path.exists(p):
            print(f"  ⚠️ 缺默认图 {img_id}，跳过 {label}")
            continue
        img = XLImage(p)
        img.width, img.height = w, h
        ws.add_image(img, f"{get_column_letter(col)}{row}")
        inserted_imgs += 1
        print(f"  插入 {label} -> {get_column_letter(col)}{row} ({w}x{h}px)")

    wb.save(DST)
    print(f"\n✅ 净化模板已生成: {DST}")
    print(f"   清除公式: {cleared_formula} 个 | 清除 DISPIMG: {cleared_dispimg} 个 | 插入图片: {inserted_imgs} 张")


if __name__ == "__main__":
    main()
