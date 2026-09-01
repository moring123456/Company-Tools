# -*- coding: utf-8 -*-
"""ui.py — 采购单生成模块的 Streamlit 界面（可被 main_app 单入口集成）

独立入口: app.py (set_page_config + title + render_purchase_order_page)
平台集成: main_app.py 菜单调用 render_purchase_order_page()
"""
import io
import os
import tempfile

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 双兼容导入：平台集成(包上下文) / 独立运行(脚本目录)
try:
    from .core.parser import parse_input
    from .core.splitter import split
    from .core.generator import generate_all, make_zip
    from .core.lingxing import build_lingxing
    _PKG = True
except ImportError:
    from core.parser import parse_input
    from core.splitter import split
    from core.generator import generate_all, make_zip
    from core.lingxing import build_lingxing
    _PKG = False

_SELF_DIR = os.path.dirname(os.path.abspath(__file__))
# 下载模板底版：用户提供的真实模板（含主表公式款号/真实示例/图片sheet含DISPIMG默认图）
INPUT_TEMPLATE_PATH = os.path.join(_SELF_DIR, "templates", "input_template.xlsx")


def _generate_input_template() -> bytes:
    """生成/返回下载模板 xlsx 字节。
    优先读取 templates/input_template.xlsx（用户提供的真实模板底版，
    含 4 个 sheet：主表/面料表/配置表/图片）；文件不存在时 fallback 动态生成基础版。"""
    if os.path.exists(INPUT_TEMPLATE_PATH):
        with open(INPUT_TEMPLATE_PATH, "rb") as f:
            return f.read()
    wb = openpyxl.Workbook()
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill('solid', fgColor='D9E1F2')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Sheet1 主表
    ws1 = wb.active
    ws1.title = '主表'
    ws1.append(['款号', 'SKU编码', 'SKU名称', '颜色', '尺码', '采购数', '工厂', '布行', '品名', '备注'])
    for cell in ws1[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
    for c, w in enumerate([10, 20, 25, 18, 8, 10, 10, 15, 12, 15], 1):
        ws1.column_dimensions[get_column_letter(c)].width = w
    
    # 示例行 - 纯色款 PPY0015 长袖上衣（工厂=媛媛，布行=宏裕布行）
    ws1.append(['PPY0015', 'PPY0015-CS-0102', 'PPY0015-CS 长袖上衣 纯色', '#225 黑色', 'M', 160, '媛媛', '宏裕布行', '牛奶丝', ''])
    ws1.append(['PPY0015', 'PPY0015-CS-0103', 'PPY0015-CS 长袖上衣 纯色', '#225 黑色', 'L', 160, '媛媛', '宏裕布行', '牛奶丝', ''])
    ws1.append(['PPY0015', 'PPY0015-CS-0104', 'PPY0015-CS 长袖上衣 纯色', '#225 黑色', 'XL', 480, '媛媛', '宏裕布行', '牛奶丝', ''])
    ws1.append(['PPY0015', 'PPY0015-CS-0105', 'PPY0015-CS 长袖上衣 纯色', '#225 黑色', '2XL', 480, '媛媛', '宏裕布行', '牛奶丝', ''])
    ws1.append(['PPY0015', 'PPY0015-CS-0106', 'PPY0015-CS 长袖上衣 纯色', '#225 黑色', '3XL', 240, '媛媛', '宏裕布行', '牛奶丝', ''])

    # 示例行 - 印花款 PPY1042 V领短袖长裙（工厂=聚进，布行=剑邑布行）
    ws1.append(['PPY1042', 'PPY1042-10-2603', 'PPY1042-YH V领短袖长裙 印花', 'YH1026 漫天三月红', 'L', 10, '聚进', '剑邑布行', '牛奶丝', ''])
    ws1.append(['PPY1042', 'PPY1042-10-2604', 'PPY1042-YH V领短袖长裙 印花', 'YH1026 漫天三月红', 'XL', 10, '聚进', '剑邑布行', '牛奶丝', ''])
    ws1.append(['PPY1042', 'PPY1042-10-2605', 'PPY1042-YH V领短袖长裙 印花', 'YH1026 漫天三月红', '2XL', 20, '聚进', '剑邑布行', '牛奶丝', ''])
    ws1.append(['PPY1042', 'PPY1042-10-2606', 'PPY1042-YH V领短袖长裙 印花', 'YH1026 漫天三月红', '3XL', 30, '聚进', '剑邑布行', '牛奶丝', ''])

    # 示例行 - 混搭色款 PPY4005 亨利领睡衣套装（颜色含+号，布行用空格分隔2个布行）
    ws1.append(['PPY4005', 'PPY4005-HD-0103', 'PPY4005-YH 亨利领睡衣套装 混搭', '#179 海军蓝+YH5126 棕蓝格子', 'L', 10, '嘉品', '宏裕布行 剑邑布行', '牛奶丝', ''])
    ws1.append(['PPY4005', 'PPY4005-HD-0104', 'PPY4005-YH 亨利领睡衣套装 混搭', '#179 海军蓝+YH5126 棕蓝格子', 'XL', 10, '嘉品', '宏裕布行 剑邑布行', '牛奶丝', ''])
    ws1.append(['PPY4005', 'PPY4005-HD-0105', 'PPY4005-YH 亨利领睡衣套装 混搭', '#179 海军蓝+YH5126 棕蓝格子', '2XL', 20, '嘉品', '宏裕布行 剑邑布行', '牛奶丝', ''])
    ws1.append(['PPY4005', 'PPY4005-HD-0106', 'PPY4005-YH 亨利领睡衣套装 混搭', '#179 海军蓝+YH5126 棕蓝格子', '3XL', 20, '嘉品', '宏裕布行 剑邑布行', '牛奶丝', ''])

    # Sheet2 面料表
    ws2 = wb.create_sheet('面料表')
    ws2.append(['款号', '产品名称', '品名', '纸样名称', '洗水唛成分', '每条布出货数(件)', '是否需要压缩袋', '包装袋规格'])
    for cell in ws2[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
    for c, w in enumerate([10, 20, 12, 30, 22, 18, 16, 14], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.append(['PPY0015', '长袖上衣-纯色 印花', '牛奶丝', 'PPY0015V领长袖（丝光棉）（2024.5.9）', '95% Rayon\n5% Spandex', 70, '是', '27.5*37.5'])
    ws2.append(['PPY1042', 'PPY1042 V领短袖长裙', '牛奶丝', 'PPY1042 连衣裙（2026.5.7）', '95% Polyester5% Spandex', 50, '是', '27.5*37.5'])
    ws2.append(['PPY4005', '长袖亨利领睡衣套装', '牛奶丝', 'PPY4005 纽扣长袖长裤睡衣（2026.5.29）', '95% Polyester5% Spandex', 40, '是', '29.5*37.5'])

    # Sheet3 配置表
    ws3 = wb.create_sheet('配置表')
    ws3.append(['配置项', '值'])
    for cell in ws3[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 40

    # 配置表示例数据
    ws3.append(['品牌', 'POPYOUNG'])
    ws3.append(['账号代码', 'MLD-AM-US-PPY'])
    ws3.append(['申请人', '肖体桥'])
    ws3.append(['联系地址', '东莞启创-朗晨'])
    ws3.append(['采购仓库', '启创-福茂仓库'])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_purchase_order_page():
    """采购单自动生成模块主界面（不含 set_page_config/title，避免与平台冲突）"""
    st.caption("朗晨 PPY 采购单批量生成 · 输入 1 文件 4 工作表（主表 / 面料表 / 配置表 / 图片）→ 输出 按款×工厂 拆分 + 每布行一张申购单")

    # ---------------- 状态初始化 ----------------
    if "uploaded_name" not in st.session_state:
        st.session_state.uploaded_name = None
    if "data" not in st.session_state:
        st.session_state.data = None
    if "edited_qty1" not in st.session_state:
        st.session_state.edited_qty1 = {}   # {sku: qty1}

    # ---------------- 步骤 1: 上传 ----------------
    st.subheader("1️⃣ 上传输入文件")
    uploaded = st.file_uploader("选择 .xlsx 输入文件", type=["xlsx"])

    # 动态生成空输入模板供下载（不依赖本地文件）
    col_up, col_dl = st.columns([3, 1])
    with col_dl:
        _tpl_bytes = _generate_input_template()
        st.download_button(
            label="📄 下载空输入模板",
            data=_tpl_bytes,
            file_name="采购单输入模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="下载空模板，按格式填写后上传即可",
        )

    if uploaded is not None and uploaded.name != st.session_state.uploaded_name:
        st.session_state.uploaded_name = uploaded.name
        st.session_state.edited_qty1 = {}
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            f.write(uploaded.getvalue())
            st.session_state._tmp_path = f.name
        st.session_state.data = parse_input(st.session_state._tmp_path)

    data = st.session_state.data
    if data is None:
        st.info("请先上传输入文件。格式要求见侧边栏说明。")
        return

    # ---------------- 步骤 2: 校验 ----------------
    st.subheader("2️⃣ 数据校验")
    if data.errors:
        st.error(f"❌ 发现 {len(data.errors)} 个问题，请修正输入文件后重新上传：")
        df_err = pd.DataFrame(data.errors, columns=["行号", "字段", "问题", "建议"])
        st.dataframe(df_err, use_container_width=True, hide_index=True)
        return
    else:
        st.success(f"✅ 校验通过：{len(data.rows)} 个 SKU / {len(data.fabrics)} 个款 / 配置表已读取")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("品牌", data.config.brand)
        c2.metric("账号代码", data.config.account_code)
        c3.metric("申请人", data.config.applicant)
        c4.metric("联系地址", data.config.office)

    # ---------------- 步骤 3: 预览 ----------------
    st.subheader("3️⃣ 拆单预览")

    # 混搭色编辑（先于 split 执行）
    mixed_skus = [(sr.sku, sr.color, sr.qty, sr.supplier, sr.row_no)
                  for sr in data.rows
                  if sr.color and "+" in str(sr.color)]
    if mixed_skus:
        st.markdown("#### ⚠️ 混搭色拆分（默认对半，可修改前半段件数）")
        df_mixed = pd.DataFrame([
            {
                "SKU编码": sku,
                "原颜色": color,
                "总件数": qty,
                "前半段件数": st.session_state.edited_qty1.get(sku, -(-qty // 2)),
                "后半段件数": qty - st.session_state.edited_qty1.get(sku, -(-qty // 2)),
                "布行(前半,后半)": supplier,
            }
            for sku, color, qty, supplier, _ in mixed_skus
        ])
        edited = st.data_editor(
            df_mixed, use_container_width=True, hide_index=True,
            column_config={"前半段件数": st.column_config.NumberColumn(min_value=0, step=1)},
            key="mixed_editor",
        )
        st.session_state.edited_qty1 = {
            row["SKU编码"]: int(row["前半段件数"])
            for _, row in edited.iterrows()
        }

    # 执行拆单
    res = split(data, st.session_state.edited_qty1 or None)

    # 🖼️ 可选: 在网页直接上传图片替换订单里的图 (不传则用默认图)
    st.markdown("---")
    st.markdown("#### 🖼️ 替换订单图片（可选，不传用默认）")
    st.caption("留空 = 用程序内置默认图；上传 = 用你的图覆盖对应位置")
    img_cols = st.columns(4)
    img_labels = ["包装袋图片", "主唛", "吊牌", "温馨提示卡片"]
    uploaded_imgs = {}
    for col, label in zip(img_cols, img_labels):
        with col:
            f = st.file_uploader(label, type=["png", "jpg", "jpeg"], key=f"img_{label}")
            if f is not None:
                uploaded_imgs[label] = f.read()

    st.markdown(f"#### 将生成 **{len(res.units)}** 份采购单文件")

    # 步骤 3.5：可编辑订单列表（用户在生成前可改订单号/纸样编号/采购日期）
    st.markdown("##### ✏️ 编辑订单信息（改后直接点下方生成即可生效）")
    st.caption("订单号 = 文件名前缀 · 纸样编号 = 订单 B4 写入值 · 采购日期 = 影响文件名+订单 H5")
    edit_rows = []
    for u in res.units:
        edit_rows.append({
            "订单号": u.order_no,
            "款号": u.style,
            "纸样编号": u.pattern,
            "工厂": u.factory,
            "布行": "、".join(g.supplier for g in u.supplier_groups),
            "采购日期": u.date_str or "",
            "采购明细": "、".join(f"{cr.color}({cr.total}件)" for cr in u.color_rows),
        })
    df_edit = pd.DataFrame(edit_rows)
    edited = st.data_editor(
        df_edit, use_container_width=True, hide_index=True,
        column_config={
            "订单号": st.column_config.TextColumn(help="改后文件名自动同步"),
            "款号": st.column_config.TextColumn(disabled=True),
            "纸样编号": st.column_config.TextColumn(help="写到订单 B4"),
            "工厂": st.column_config.TextColumn(disabled=True),
            "布行": st.column_config.TextColumn(disabled=True),
            "采购明细": st.column_config.TextColumn(disabled=True),
            "采购日期": st.column_config.TextColumn(help="改后所有订单同步更新日期段+文件名"),
        },
        key="order_editor",
    )
    # 应用编辑到 res.units
    for i, u in enumerate(res.units):
        row = edited.iloc[i]
        new_date = str(row["采购日期"]).strip() if row["采购日期"] else ""
        if new_date and new_date != (u.date_str or ""):
            old_date = u.date_str or ""
            if old_date and old_date in u.order_no:
                u.order_no = u.order_no.replace(old_date, new_date, 1)
            else:
                # 订单号里没旧日期（极端情况），按 prefix+date+seq 重组
                import re as _re
                m = _re.match(r"^([A-Za-z]*)(\d{8})(\d{2,})$", u.order_no)
                if m:
                    u.order_no = f"{m.group(1)}{new_date}{m.group(3)}"
            u.date_str = new_date
        # 订单号/纸样编号
        if row["订单号"] and row["订单号"] != u.order_no:
            u.order_no = str(row["订单号"])
        if row["纸样编号"] != u.pattern:
            u.pattern = str(row["纸样编号"])
        # 文件名始终重新生成（按当前 order_no + style + product_name + type + factory）
        u.file_name = (
            f"{u.order_no}-{u.style} {u.product_name} {u.type_str} -{u.factory}.xlsx"
        ).replace("  ", " ").strip()

    preview_rows = []
    for u in res.units:
        preview_rows.append({
            "订单号": u.order_no,
            "款号": u.style,
            "产品名称": u.product_name,
            "类型": u.type_str,
            "工厂": u.factory,
            "涉及布行": "、".join(g.supplier for g in u.supplier_groups),
            "颜色数": len(u.color_rows),
            "总件数": u.total_qty,
            "生成文件名": u.file_name,
        })
    st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

    with st.expander("🔍 查看各文件明细（订单颜色行 + 申购单）"):
        for u in res.units:
            st.markdown(f"**{u.order_no} | {u.style} | {u.factory} | {u.file_name}**")
            c_left, c_right = st.columns(2)
            with c_left:
                st.markdown("**订单颜色行**")
                st.dataframe(pd.DataFrame([
                    {"颜色": cr.color, "尺码数量": str(cr.qty_by_size), "比例": cr.ratio,
                     "总计": cr.total, "布料条数": cr.rolls, "备注": cr.remark}
                    for cr in u.color_rows
                ]), use_container_width=True, hide_index=True)
            with c_right:
                for g in u.supplier_groups:
                    st.markdown(f"**布料申购单-{g.supplier}**（合计 {g.total_rolls} 条）")
                    st.dataframe(pd.DataFrame([
                        {"颜色&色号": r.color, "数量(条)": r.rolls, "布行": r.supplier}
                        for r in g.rows
                    ]), use_container_width=True, hide_index=True)

    # ---------------- 步骤 4: 生成 ----------------
    st.subheader("4️⃣ 生成采购单")
    if st.button("⚡ 一键生成全部文件", type="primary"):
        with st.spinner("正在生成…"):
            out_dir = tempfile.mkdtemp(prefix="purchase_out_")
            # 网页上传的图片优先于输入文件「图片」sheet
            if uploaded_imgs:
                data.images = {**data.images, **uploaded_imgs}
            generate_all(data, res, out_dir)
            # 领星批量导入采购单（模板底版 + 标识号分组）
            try:
                build_lingxing(data, res, out_dir)
            except Exception as e:
                st.warning(f"领星文件生成失败（不影响采购单）：{e}")
            date_str = res.units[0].order_no[3:11] if res.units else ""
            zip_name = f"{os.path.splitext(st.session_state.uploaded_name)[0]}_采购单.zip"
            zip_path = os.path.join(tempfile.gettempdir(), zip_name)
            make_zip(out_dir, zip_path, date_str)
            with open(zip_path, "rb") as f:
                zip_bytes = f.read()
        st.success(f"✅ 已生成 {len(res.units)} 份采购单 + 汇总明细表 + 领星导入文件，打包完成！")
        st.download_button(
            label=f"📥 下载 {zip_name}（{len(zip_bytes)/1024:.0f} KB）",
            data=zip_bytes, file_name=zip_name, mime="application/zip", type="primary",
        )

    # ---------------- 侧边栏说明 ----------------
    with st.sidebar:
        st.markdown("### 📋 输入格式说明")
        st.markdown("**Sheet1 主表**：款号 / SKU编码 / SKU名称 / 颜色 / 尺码 / 采购数 / 工厂 / 布行 / 品名 / 备注")
        st.markdown("**Sheet2 面料表**：款号 / 产品名称 / 品名 / 纸样名称 / 洗水唛成分 / 每条布出货数(件) / 是否需要压缩袋 / 包装袋规格")
        st.markdown("**Sheet3 配置表**：品牌 / 账号代码 / 申请人 / 联系地址 / 采购仓库（下拉）")
        st.divider()
        st.markdown("### ⚙️ 规则要点")
        st.markdown("- 拆单：按 **款号×工厂** 一份文件；订单号 = PPY+日期+序号")
        st.markdown("- 布料条数 = 件数 ÷ 每条布出货数（向上取整）")
        st.markdown("- 混搭色：布行填 2 个（逗号/空格/分号均可），拆分默认对半可改")
        st.markdown("- 同款同色 SKU 的工厂/布行必须一致（防复制错误）")
        st.markdown("- 输出含 领星批量导入采购单（供应商自动加\"工厂\"后缀）")
