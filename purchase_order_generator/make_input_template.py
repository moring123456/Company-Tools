# -*- coding: utf-8 -*-
"""生成输入模板: PPY采购-输入模板.xlsx（3 工作表，带表头样式/示例/下拉）"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"D:\WorkBuddy-存储\2026-06-24-11-13-53\purchase-order-generator\output\PPY采购-输入模板.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="FFB03A2E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

# ============ Sheet1 主表 ============
ws = wb.active
ws.title = "主表"
headers1 = ["款号", "SKU编码", "SKU名称", "颜色", "尺码", "采购数", "工厂", "布行", "品名", "备注"]
ws.append(headers1)
sample_rows = [
    ["PPY4005", "PPY4005-YH-2143", "PPY4005-YH 亨利领睡衣套装 印花", "HX214 雪花红底", "L", 10, "嘉品", "剑邑布行", "牛奶丝", "新颜色"],
    ["PPY4005", "PPY4005-HD-0103", "PPY4005-YH 亨利领睡衣套装 混搭", "#179 海军蓝+YH5126 棕蓝格子", "L", 10, "嘉品", "剑邑布行,宏裕布行", "牛奶丝", "混搭色布行填2个: 逗号/空格/分号均可"],
]
for r in sample_rows:
    ws.append(r)
# 表头样式
for c in range(1, 11):
    cell = ws.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BOX
# 列宽
widths = [12, 22, 34, 26, 8, 9, 10, 22, 12, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
# 尺码下拉 S~6XL
dv_size = DataValidation(type="list", formula1='"S,M,L,XL,2XL,3XL,4XL,5XL,6XL"', allow_blank=True)
ws.add_data_validation(dv_size)
dv_size.add("E2:E500")
# 示例行提示
ws["A3"].comment = None

# ============ Sheet2 面料表 ============
ws2 = wb.create_sheet("面料表")
headers2 = ["款号", "产品名称", "品名", "纸样名称", "洗水唛成分", "每条布出货数(件)", "是否需要压缩袋", "包装袋规格"]
ws2.append(headers2)
ws2.append(["PPY4005", "长袖亨利领睡衣套装", "牛奶丝", "PPY4005(2023.7.25)", "95% Rayon\n5% Spandex", 40, "是", "29.5*37.5"])
for c in range(1, 9):
    cell = ws2.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BOX
for i, w in enumerate([12, 22, 14, 22, 24, 16, 14, 14], 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
dv_bag = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
ws2.add_data_validation(dv_bag)
dv_bag.add("G2:G500")

# ============ Sheet3 配置表 ============
ws3 = wb.create_sheet("配置表")
config_rows = [["品牌", "POPYOUNG"], ["账号代码", "MLD-AM-US-PPY"], ["申请人", "肖体桥"],
               ["联系地址", "东莞启创-朗晨"], ["采购仓库", "启创-立祥仓库B区"]]
for r in config_rows:
    ws3.append(r)
for c in range(1, 3):
    cell = ws3.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 26
# 采购仓库下拉（领星仓库枚举）
dv_wh = DataValidation(type="list", formula1='"启创-利得仓库,启创-福茂仓库,启创-立祥仓库A区,启创-立祥仓库B区"',
                       allow_blank=True)
ws3.add_data_validation(dv_wh)
dv_wh.add("B5:B500")

wb.save(OUT)
print("输入模板已生成:", OUT)
